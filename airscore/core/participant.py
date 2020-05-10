"""
Participant Library

contains
    Participant class
    function to import participants from excel file

Use: from participant import Participant

Stuart Mackintosh Antonio Golfari - 2019
"""

from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import aliased

from calcUtils import get_date
from civlrankings import create_participant_from_CIVLID, create_participant_from_name
from db_tables import TblParticipant
from logger import Logger
from myconn import Database


class Participant(object):
    """Participant definition, DB operations
    """

    def __init__(self, par_id=None, comp_id=None, ID=None, civl_id=None, name=None, sex=None, birthdate=None,
                 nat=None, glider=None, glider_cert=None, sponsor=None, fai_id=None, fai_valid=1, xcontest_id=None,
                 team=None, nat_team=1, status=None, ranking=None, paid=None, pil_id=None):

        self.par_id = par_id  # pil_id
        self.comp_id = comp_id  # comp_id
        self.ID = ID  # int
        self.civl_id = civl_id  # int
        self.name = name  # str
        self.sex = sex  # 'M', 'F'
        self.birthdate = birthdate  # in datetime.date (Y-m-d) format
        self.nat = nat  # str
        self.glider = glider  # str
        self.glider_cert = glider_cert  # str
        self.sponsor = sponsor  # str
        self.fai_id = fai_id  # str
        self.fai_valid = fai_valid  # bool
        self.xcontest_id = xcontest_id  # str
        self.team = team  # ?
        self.nat_team = nat_team  # default 1
        self.paid = paid  # bool
        self.status = status  # 'confirmed', 'waiting list', 'wild card', 'cancelled', ?
        self.ranking = ranking  # WPRS Ranking?
        self.live_id = None  # int
        self.pil_id = pil_id  # PilotView id

    def __setattr__(self, attr, value):
        if attr in ('name', 'glider') and type(value) is str:
            value = value.title()
        elif attr in ('nat', 'sex') and type(value) is str:
            value = value.upper()
        self.__dict__[attr] = value

    @property
    def pilot_birthdate_str(self):
        return self.birthdate.strftime("%Y-%m-%d")

    @property
    def female(self):
        return 1 if self.sex == 'F' else 0

    def as_dict(self):
        return self.__dict__

    def __str__(self):
        out = ''
        out += 'Pilot:'
        out += f'{self.name} - CIVL_ID {self.civl_id} \n'
        out += f'{self.glider}  | {self.sponsor}\n'
        return out

    @staticmethod
    def read(par_id):
        """Reads pilot registration from database
        takes pil_id as argument"""
        from db_tables import TblParticipant as R

        if not (type(par_id) is int and par_id > 0):
            print(f"par_id needs to be int > 0, {par_id} was given")
            return None

        pilot = Participant(par_id=par_id)

        with Database() as db:
            # get pilot details.
            q = db.session.query(R).get(par_id)
            db.populate_obj(pilot, q)
        return pilot

    @staticmethod
    def from_dict(d):
        par = Participant()
        for key, value in d.items():
            if hasattr(par, key):
                setattr(par, key, value)
        return par

    def to_db(self, session=None):
        """stores or updates Participant to AirScore database"""

        with Database(session) as db:
            try:
                if not self.par_id:
                    pil = TblParticipant()
                    db.session.add(pil)
                    db.session.flush()
                    self.par_id = pil.par_id
                else:
                    pil = db.session.query(TblParticipant).get(self.par_id)

                for attr in dir(pil):
                    if not attr[0] == '_' and hasattr(self, attr):
                        setattr(pil, attr, getattr(self, attr))
                db.session.flush()

            except SQLAlchemyError:
                print('Participant storing error')
                db.session.rollback()
                return None

        return self.par_id

    @staticmethod
    def from_fsdb(pil, from_CIVL=0):
        """gets pilot obj. from FSDB file
            Input:
                - pil:          lxml.etree: FsParticipant section
                - from_CIVL:    BOOL: look for pilot on CIVL database"""

        CIVLID = None if not pil.get('CIVLID') else int(pil.get('CIVLID'))
        name = pil.get('name')
        # print(CIVLID, name)
        pilot = None
        if from_CIVL:
            '''check CIVL database'''
            if CIVLID:
                pilot = create_participant_from_CIVLID(CIVLID)
            else:
                pilot = create_participant_from_name(name)
        '''check if we have a result and name is similar'''
        if not (pilot and any(n in pilot.name for n in name)):
            '''get all pilot info from fsdb file'''
            if from_CIVL:
                print('*** no result in CIVL database, getting data from FSDB file')
            pilot = Participant(name=name, civl_id=CIVLID)
            pilot.sex = 'F' if int(pil.get('female') if pil.get('female') else 0) > 0 else 'M'
            pilot.nat = pil.get('nat_code_3166_a3')

        pilot.birthdate = get_date(pil.get('birthday'))
        pilot.ID = int(pil.get('id'))
        pilot.glider = pil.get('glider')
        pilot.sponsor = pil.get('sponsor')
        """check fai is int"""
        if pil.get('fai_licence'):
            pilot.fai_valid = True
            pilot.fai_id = pil.get('fai_licence')
        else:
            pilot.fai_valid = False
            pilot.fai_id = None
        """check Live ID"""
        node = pil.find('FsCustomAttributes')
        if node is not None:
            childs = node.findall('FsCustomAttribute')
            live = next(el for el in childs if el.get('name') == 'Live')
            if live is not None:
                pilot.live_id = int(live.get('value'))
                print(pilot.live_id)

        return pilot

    @staticmethod
    def from_profile(pilot_id, comp_id=None, session=None):
        from db_tables import PilotView, TblCountryCode
        with Database(session) as db:
            try:
                result = db.session.query(PilotView).get(pilot_id)
                if result:
                    pilot = Participant(pil_id=pilot_id, comp_id=comp_id)
                    pilot.name = ' '.join([result.first_name.title(), result.last_name.title()])
                    pilot.glider = ' '.join([result.glider_brand.title(), result.glider])
                    c = aliased(TblCountryCode)
                    pilot.nat = db.session.query(c.natIso3).filter(c.natId == result.nat).scalar()
                    for attr in ['sex', 'civl_id', 'fai_id', 'sponsor', 'xcontest_id', 'glider_cert']:
                        if hasattr(result, attr):
                            setattr(pilot, attr, getattr(result, attr))
                    return pilot
                else:
                    print(f'Error: No result has been found for profile id {pilot_id}')
                    return None
            except SQLAlchemyError as e:
                error = str(e.__dict__)
                print(f"Error storing result to database")
                db.session.rollback()
                db.session.close()
                return error


def extract_participants_from_excel(comp_id, filename, from_CIVL=False):
    """Gets participants from external file (Airtribune Participants list in Excel format;
    Returns a list of Participant objects
    Input:
        comp_id:    INT comp_id
        filename:   STR filename
        from_CIVL:  BOOL retrieve data from CIVL database using CIVLID
    - read Airtribune XSLX file

    excel file format:
    column name on row 1
    columns:
    id,name,nat,female,birthday,glider,color,sponsor,fai_licence,CIVILID,club,team,class,Live(optional)
    """

    '''create logging and disable output'''
    # Logger('ON', 'import_participants.txt')

    print(f"Comp ID: {comp_id} | filename: {filename}")

    '''load excel file'''
    try:
        workbook = load_workbook(filename=filename)
    except:
        print('excel file importing error')
        return

    '''active sheet'''
    sheet = workbook.active

    '''check validity'''
    if not (sheet['A1'].value == 'id'):
        exit()

    pilots = []
    for row in sheet.iter_rows(min_row=2,
                               min_col=1,
                               max_col=14,
                               values_only=True):
        if not row[0]:
            'EOF'
            break

        ''' Check CIVL database if active'''
        pil = None
        if from_CIVL and row[9]:
            pil = create_participant_from_CIVLID(row[9])
        if pil is None:
            pil = Participant(civl_id=row[9], name=row[1], nat=row[2], sex='F' if row[3] == 1 else 'M')
        pil.ID = row[0]
        pil.comp_id = comp_id
        pil.birthdate = None if row[4] is None else row[4].date()  # row[4] should be datetime
        pil.glider = row[5]
        pil.sponsor = row[7]
        pil.team = row[11]
        pil.fai_id = row[8]
        pil.fai_valid = 0 if row[8] is None else 1
        pil.live_id = row[13]
        pilots.append(pil)

    # ''' now restore stdout function '''
    # Logger('OFF')

    return pilots


def register_from_profiles_list(comp_id, pilots):
    """ gets comp_id and pil_id list
        registers pilots to comp"""
    if not (comp_id and pilots):
        print(f"error: comp_id does not exist or pilots list is empty")
        return None
    participants = []
    with Database() as db:
        for pilot in pilots:
            participants.append(Participant.from_profile(pilot, comp_id, db.session))
        mass_import_participants(comp_id, participants, db.session)
    return True


def unregister_from_profiles_list(comp_id, pilots):
    """ takes comp_id and list of pil_id
        unregisters pilots from comp"""
    from sqlalchemy import and_
    if not (comp_id and pilots):
        print(f"error: comp_id does not exist or pilots list is empty")
        return None
    with Database() as db:
        try:
            results = db.session.query(TblParticipant).filter(and_(TblParticipant.comp_id == comp_id,
                                                                   TblParticipant.pil_id.in_(pilots)))
            results.delete(synchronize_session=False)
        except SQLAlchemyError as e:
            error = str(e.__dict__)
            print(f"Error deleting participants from database: {error}")
            db.session.rollback()
            db.session.close()
            return error
    return True


def unregister_participant(comp_id, par_id):
    """ takes comp_id and a par_id
        unregisters participant from comp.
        in reality we don't need compid but it is a safeguard"""
    from sqlalchemy import and_

    with Database() as db:
        try:
            results = db.session.query(TblParticipant).filter(and_(TblParticipant.comp_id == comp_id,
                                                                   TblParticipant.par_id == par_id))
            results.delete(synchronize_session=False)
        except SQLAlchemyError as e:
            error = str(e.__dict__)
            print(f"Error deleting participants from database: {error}")
            db.session.rollback()
            db.session.close()
            return error
    return True


def unregister_all_exteranl_participants(comp_id):
    """ takes comp_id and unregisters all participants from comp without a pil_id."""
    from sqlalchemy import and_

    with Database() as db:
        try:
            results = db.session.query(TblParticipant).filter(and_(TblParticipant.comp_id == comp_id,
                                                                   TblParticipant.pil_id == None))
            results.delete(synchronize_session=False)
        except SQLAlchemyError as e:
            error = str(e.__dict__)
            print(f"Error deleting participants from database: {error}")
            db.session.rollback()
            db.session.close()
            return error
    return True


def mass_unregister(pilots):
    """ gets par_id list
        unregisters pilots from comp"""
    if not pilots:
        print(f"error: pilots list is empty")
        return None
    with Database() as db:
        try:
            results = db.session.query(TblParticipant).filter(TblParticipant.par_id.in_(pilots))
            results.delete(synchronize_session=False)
        except SQLAlchemyError as e:
            error = str(e.__dict__)
            print(f"Error deleting participants to database: {error}")
            db.session.rollback()
            db.session.close()
            return error
    return True


def mass_import_participants(comp_id, participants, session=None):
    """get participants to update from the list"""
    # TODO check if we already have participants for the comp before inserting, and manage update instead

    insert_mappings = []
    update_mappings = []
    for par in participants:
        r = dict(comp_id=comp_id, par_id=par.par_id)
        for key in [col for col in TblParticipant.__table__.columns.keys() if col not in r.keys()]:
            if hasattr(par, key):
                r[key] = getattr(par, key)
        if r['par_id']:
            update_mappings.append(r)
        else:
            insert_mappings.append(r)
    '''update database'''
    with Database(session) as db:
        try:
            if insert_mappings:
                db.session.bulk_insert_mappings(TblParticipant, insert_mappings)
                db.session.flush()
                for elem in insert_mappings:
                    next(par for par in participants if par.name == elem['name']).par_id = elem['par_id']
            if update_mappings:
                db.session.bulk_update_mappings(TblParticipant, update_mappings)
            db.session.commit()
        except SQLAlchemyError as e:
            error = str(e.__dict__)
            print(f"Error storing participants to database: {error}")
            db.session.rollback()
            db.session.close()
            return error
    return True
