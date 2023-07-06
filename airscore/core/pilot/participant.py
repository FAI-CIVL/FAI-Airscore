"""
Participant Library

contains
    Participant class
    function to import participants from excel file

Use: from participant import Participant

Stuart Mackintosh Antonio Golfari - 2019
"""

import datetime

from calcUtils import get_date
from db.conn import db_session
from db.tables import TblParticipant as P, TblParticipantMeta as PA, TblCompAttribute as CA
from pilot.pilot import Pilot
from sources.civlrankings import (
    create_participant_from_CIVLID,
    create_participant_from_name,
)
from sqlalchemy.orm import aliased


class Participant(Pilot):
    """Participant definition, DB operations"""

    def __init__(
        self,
        par_id=None,
        comp_id=None,
        ID=None,
        glider=None,
        glider_cert=None,
        sponsor=None,
        team=None,
        nat_team=1,
        status=None,
        paid=None,
        live_id=None,
        **kwargs,
    ):

        self.par_id = par_id  # pil_id
        self.comp_id = comp_id  # comp_id
        self.ID = ID  # int
        self.glider = glider  # str
        self.glider_cert = glider_cert  # str
        self.sponsor = sponsor  # str
        self.team = team  # ?
        self.nat_team = nat_team  # default 1
        self.paid = paid  # bool
        self.status = status  # 'confirmed', 'waiting list', 'wild card', 'cancelled', ?
        self.live_id = live_id  # int
        self.custom = {}  # dict attr_id: meta_value
        super().__init__(**kwargs)

    def __setattr__(self, attr, value):
        property_names = [p for p in dir(Participant) if isinstance(getattr(Participant, p), property)]
        if attr == 'glider' and type(value) is str:
            self.__dict__[attr] = value.title()
        elif attr in ('nat', 'sex') and type(value) is str:
            self.__dict__[attr] = value.upper()
        elif attr not in property_names:
            self.__dict__[attr] = value

    def as_dict(self):
        return self.__dict__

    def __str__(self):
        out = ''
        out += 'Participant:'
        out += f'{self.ID if self.ID else None} {self.name} - CIVL_ID {self.civl_id} \n'
        out += f'{self.glider}  | {self.sponsor}\n'
        return out

    @staticmethod
    def read(par_id: int):
        """Reads pilot registration from database
        takes pil_id as argument"""
        with db_session() as db:
            # get pilot details.
            q = db.query(P).get(par_id)
            if q:
                participant = Participant(par_id=par_id)
                q.populate(participant)
                comp_attr = db.query(CA).filter_by(comp_id=participant.comp_id, attr_key='meta')
                custom_attr = db.query(PA).filter_by(par_id=par_id)
                participant.custom = {el.attr_id: next((x.meta_value for x in custom_attr
                                                        if x.attr_id == el.attr_id), None) for el in comp_attr}
                return participant
        return None

    @staticmethod
    def from_dict(d: dict):
        """creates a Participant obj from a dictionary obj"""
        par = Participant()
        for key, value in d.items():
            if hasattr(par, key):
                setattr(par, key, value)
        return par

    def to_db(self):
        """stores or updates Participant to AirScore database"""
        if not self.par_id:
            row = P.from_obj(self)
            row.save()
            self.par_id = row.par_id
        else:
            row = P.get_by_id(self.par_id)
            row.update(**self.as_dict())
            PA.delete_all(par_id=self.par_id)
        attr = []
        for key in [k for k in self.custom.keys() if self.custom[k] is not None]:
            attr.append(PA(par_id=self.par_id, attr_id=key, meta_value=self.custom[key]))
        PA.bulk_create(attr)
        return self.par_id

    @staticmethod
    def from_fsdb(pil, from_CIVL=False, comp_attributes=None):
        """gets pilot obj. from FSDB file
        Input:
            - pil:          lxml.etree: FsParticipant section
            - from_CIVL:    BOOL: look for pilot on CIVL database"""
        from calcUtils import get_int

        CIVLID = get_int(pil.get('CIVLID')) or None
        name = pil.get('name')
        # print(CIVLID, name)
        pilot = None
        if from_CIVL:
            '''check CIVL database'''
            if CIVLID:
                pilot = create_participant_from_CIVLID(CIVLID)
            else:
                pilot = create_participant_from_name(name)
        '''check if we have a result and name is similar'''
        if not (pilot and any(n in pilot.name for n in name)):
            '''get all pilot info from fsdb file'''
            if from_CIVL:
                print('*** no result in CIVL database, getting data from FSDB file')
            pilot = Participant(name=abbreviate(name), civl_id=CIVLID)
            pilot.sex = 'F' if int(pil.get('female') if pil.get('female') else 0) > 0 else 'M'
            pilot.nat = pil.get('nat_code_3166_a3') or None
        bd = get_date(pil.get('birthday') or None)
        pilot.birthdate = None if not isinstance(bd, datetime.date) else bd
        pilot.ID = get_int(pil.get('id'))
        pilot.glider = abbreviate(pil.get('glider')) or None
        pilot.sponsor = abbreviate(pil.get('sponsor')) or None
        '''check fai is int'''
        if pil.get('fai_licence') in (0, '0', '', None):
            pilot.fai_valid = False
            pilot.fai_id = None
        else:
            pilot.fai_valid = True
            pilot.fai_id = None if pil.get('fai_licence') == '1' else pil.get('fai_licence')
        '''check custom attributes'''
        node = pil.find('FsCustomAttributes')
        if node is not None:
            childs = node.findall('FsCustomAttribute')
            pilot.attributes = []
            for el in childs:
                if el.get('value') in (None, ''):
                    continue
                elif el.get('name').lower() == 'live' and len(el.get('value')) < 11 and el.get('value').isdigit():
                    pilot.live_id = int(el.get('value'))
                elif el.get('name').lower() == 'team':
                    pilot.team = el.get('value')
                elif el.get('name').lower() in ('fai_id', 'fai_licence'):
                    pilot.fai_id = el.get('value')
                else:
                    pilot.attributes.append({'attr_value': el.get('name'), 'meta_value': el.get('value')})

        return pilot

    @staticmethod
    def from_profile(pilot_id: int, comp_id: int):
        """creates a Participant obj from internal PilotView database table"""
        from db.tables import PilotView, TblCountryCode

        with db_session() as db:
            result = db.query(PilotView).get(pilot_id)
            if result:
                try:
                    pilot = Participant(pil_id=pilot_id, comp_id=comp_id)
                    pilot.name = abbreviate(' '.join([result.first_name, result.last_name]))
                    if result.glider_brand:
                        pilot.glider = ' '.join([result.glider_brand.title(), result.glider])
                    else:
                        pilot.glider = result.glider
                    c = aliased(TblCountryCode)
                    pilot.nat = db.query(c.natIoc).filter(c.natId == result.nat).scalar()
                    for attr in ['sex', 'civl_id', 'fai_id', 'sponsor', 'xcontest_id', 'glider_cert']:
                        if hasattr(result, attr):
                            setattr(pilot, attr, getattr(result, attr))
                    return pilot
                except AttributeError:
                    print(f'Error: pilot id {pilot_id} ha no name or no last name')
            else:
                print(f'Error: No result has been found for profile id {pilot_id}')
                return None


def extract_participants_from_excel(comp_id: int, filename, certs: list, from_CIVL=False):
    """Gets participants from external file (Airtribune Participants list in Excel format;
    Returns a list of Participant objects
    Input:
        comp_id:    INT comp_id
        filename:   STR filename
        from_CIVL:  BOOL retrieve data from CIVL database using CIVLID
    - read Airtribune XLSX file

    excel file format:
    column name on row 1
    columns:
    id,name,nat,female,birthday,glider,color,sponsor,fai_licence,CIVILID,club,team,class,Live(optional)
    """
    from openpyxl import load_workbook
    from ranking import CompAttribute
    from openpyxl.utils.exceptions import InvalidFileException

    '''load excel file'''
    try:
        workbook = load_workbook(filename=filename)
    except FileNotFoundError:
        print('excel file not found')
        return
    except InvalidFileException:
        print('excel file importing error')
        return
    '''active sheet'''
    sheet = workbook.active
    '''check validity'''
    if not (sheet['A1'].value == 'id'):
        print('excel file does not seem to be the correct template.')
        return
    pilots = []

    '''custom attributes'''
    col = 15
    custom_attributes = []
    '''check if class column has correct glider certification, otherwise add a class custom attribute'''
    classes = set(r[0] for r in sheet.iter_rows(min_row=2, min_col=13, max_col=13, values_only=True) if r[0])
    classes_are_certs = all(c in certs for c in classes)
    if not classes_are_certs and len(classes) > 0:
        custom_attributes.append(CompAttribute(comp_id=comp_id, attr_key='meta', attr_value='class'))
    while True:
        value = sheet.cell(1, col).value
        if not value:
            'no custom attributes'
            break
        custom_attributes.append(CompAttribute(comp_id=comp_id, attr_key='meta', attr_value=value))
        col += 1

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=col-1, values_only=True):
        if not row[0]:
            'EOF'
            break
        ''' Check CIVL database if active'''
        pil = None
        if from_CIVL and row[9]:
            pil = create_participant_from_CIVLID(row[9])
        if pil is None:
            pil = Participant(civl_id=row[9], name=abbreviate(row[1]), nat=row[2], sex='F' if row[3] == 1 else 'M')
        if custom_attributes:
            pil.attributes = []
        pil.ID = row[0]
        pil.comp_id = comp_id
        bd = get_date(row[4] or None)
        pil.birthdate = None if not isinstance(bd, datetime.date) else bd
        pil.glider = abbreviate(row[5]) or None
        if classes_are_certs:
            pil.glider_cert = row[12]
        elif len(classes) > 0:
            pil.attributes.append({'attr_value': 'class', 'meta_value': row[12]})
        pil.sponsor = abbreviate(row[7]) or None
        pil.team = row[11]
        pil.fai_id = row[8]
        pil.fai_valid = 0 if row[8] is None else 1
        pil.live_id = row[13]
        '''custom attributes'''
        if custom_attributes:
            # pil.attributes = []
            for idx, el in enumerate(c for c in custom_attributes if not c.attr_value == 'class'):
                pil.attributes.append({'attr_value': el.attr_value, 'meta_value': row[14+idx]})
        pilots.append(pil)

    return pilots, custom_attributes


def register_from_profiles_list(comp_id: int, pilots_ids: list):
    """gets comp_id and pil_id list
    registers pilots to comp"""
    if not (comp_id and pilots_ids):
        print(f"error: comp_id does not exist or pilots list is empty")
        return None
    participants = []
    for pil_id in pilots_ids:
        participants.append(Participant.from_profile(pil_id, comp_id))
    participants = get_valid_ids(comp_id, participants)
    mass_import_participants(comp_id, participants, check_ids=False)
    return True


def unregister_from_profiles_list(comp_id: int, pilots):
    """takes comp_id and list of pil_id
    unregisters pilots from comp"""
    from sqlalchemy import and_

    if not (comp_id and pilots):
        print(f"error: comp_id does not exist or pilots list is empty")
        return None
    with db_session() as db:
        results = db.query(P).filter(and_(P.comp_id == comp_id, P.pil_id.in_(pilots)))
        results.delete(synchronize_session=False)
    return True


def unregister_participant(comp_id: int, par_id: int):
    """takes comp_id and a par_id
    unregisters participant from comp.
    in reality we don't need compid but it is a safeguard"""
    with db_session() as db:
        db.query(P).filter_by(comp_id=comp_id, par_id=par_id).delete(synchronize_session=False)
    return True


def unregister_all_external_participants(comp_id: int):
    """ takes comp_id and unregisters all participants from comp without a pil_id."""
    from sqlalchemy import and_

    try:
        with db_session() as db:
            results = db.query(P).filter(and_(P.comp_id == comp_id, P.pil_id.is_(None)))
            results.delete(synchronize_session=False)
        return True
    except Exception:
        print(f'sqlalchemy error deleting all external pilots')
        return False


def mass_unregister(pilots):
    """gets par_id list
    unregisters pilots from comp"""
    if not pilots:
        print(f"error: pilots list is empty")
        return None
    with db_session() as db:
        db.query(P).filter(P.par_id.in_(pilots)).delete(synchronize_session=False)
    return True


def unregister_all(comp_id: int):
    """gets comp_id
    unregisters all registered pilots from comp"""

    with db_session() as db:
        db.query(P).filter_by(comp_id=comp_id).delete(synchronize_session=False)
    return True


def mass_import_participants(comp_id: int, participants: list, check_ids=True):
    """get participants to update from the list
    Before inserting rows without par_id, we need to check if pilot is already in participants
    Will create a list of dicts from database, if not given as parameter"""

    objects = []
    existing = [p for p in participants if p.par_id is not None]
    for par in participants:
        row = P.from_obj(par)
        row.comp_id = comp_id
        objects.append(row)
    '''update database'''
    with db_session() as db:
        db.bulk_save_objects(objects=objects, return_defaults=True)
        db.flush()
        for idx, pil in enumerate(participants):
            if pil.par_id is None and objects[idx].par_id is not None:
                pil.par_id = objects[idx].par_id
        '''update custom attributes'''
        attr = []
        par_attr_list = [el for el in participants if any(k for k, v in el.custom.items() if v is not None)]
        if existing:
            db.query(PA).filter(PA.par_id.in_([p.par_id for p in existing])).delete(synchronize_session=False)
            db.flush()
        for el in [p for p in par_attr_list]:
            attr.extend([PA(par_id=el.par_id, attr_id=k, meta_value=v) for k, v in el.custom.items()
                         if v is not None and k is not None])
        if attr:
            db.bulk_save_objects(objects=attr)
    return True


def assign_id(comp_id: int, given_id: int = None, participants: list = None, assigned_ids: list = None) -> int:
    """assigns pilots and ID if not given and if not unique
    comp_id: comp_id
    given_id: ID that was given if any
    participants: list of participants dicts"""
    from calcUtils import get_int

    if not participants:
        participants = P.get_dicts(comp_id)
    assigned_ids = [el['ID'] for el in participants] + (assigned_ids or [])
    given_id = get_int(given_id) or None  # returns int or None
    if not given_id or not (0 < given_id < 99999) or given_id in assigned_ids:
        given_id = 101
        while True:
            if given_id in assigned_ids:
                given_id += 1
            else:
                break
    return given_id


def get_valid_ids(comp_id: int, participants: list) -> list:
    """gets a list of pilots and checks their ID validity against registered pilots and correct formats
    returns a list of pilots with correct IDs"""
    '''get participants already in competition, to avoid same id'''
    registered = P.get_dicts(comp_id)
    assigned_ids = []
    for p in participants:
        p.ID = assign_id(comp_id, p.ID, registered, assigned_ids)
        assigned_ids.append(p.ID)
    return participants


def abbreviate(string: str, length: int = 100) -> str:
    """function to format participant string attributes that could be longer than database field.
    it tries to shorten string parts starting from last word, replacing with dotted initial letter, until string length
    is within given limit.
    If fails, returns the given string truncated within given length limit."""

    if not isinstance(string, str):
        return ''
    if len(string) > length:
        data = string.split()
        data.reverse()
        while len(string) > length:
            try:
                idx = data.index(next(el for el in data if len(el) > 2))
                data[idx] = data[idx][0] + '.'
                string = ' '.join(reversed(data))
                print(f'name: {string}')
            except (ValueError, StopIteration):
                idx = 99
                while string[idx] != ' ':
                    idx -= 1
                string = string[:idx]
    return string


def get_attributes(par_id, attr_list: list = None) -> dict:
    if not attr_list:
        comp_id = P.get_by_id(par_id).comp_id
        attr_list = CA.get_all(comp_id=comp_id, attr_key='meta')
    rows = PA.get_all(par_id=par_id)
    custom = {}
    for el in attr_list:
        val = next((x.meta_value for x in rows if x.attr_id == el.attr_id), None)
        custom[el.attr_id] = val
    return custom


def get_participants_meta(par_ids: list) -> list:
    with db_session() as db:
        custom_list = db.query(PA).filter(PA.par_id.in_(par_ids), PA.attr_id).order_by(PA.par_id).all()
        return custom_list
